import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy

# Load the workbook and the specific sheet
wb = load_workbook('sample_formats3.xlsx')
ws = wb.active

# Read the data into a DataFrame
df_in = pd.read_excel('sample_formats3.xlsx')

# Prepare a new workbook to write with formats
new_wb = Workbook()
new_ws = new_wb.active

# Transfer data and styles
for r_idx, row in enumerate(dataframe_to_rows(df_in, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        new_cell = new_ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Copy style from old file if needed
        old_cell = ws.cell(row=r_idx, column=c_idx)
        if old_cell.has_style:
            new_cell.font = copy(old_cell.font)
            new_cell.border = copy(old_cell.border)
            new_cell.fill = copy(old_cell.fill)
            new_cell.number_format = old_cell.number_format
            new_cell.protection = copy(old_cell.protection)
            new_cell.alignment = copy(old_cell.alignment)

# Save the new workbook
new_wb.save('formatted_output.xlsx')
